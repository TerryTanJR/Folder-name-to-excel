import os
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

# target folders' full path (if all the folders are put in a file called "Example")
# note that the target folders all need to be put into a folder
folder_path = r"C:\your own path\Example" # replace it to tour own one

# initialize Excel workbook
wb = Workbook()
ws = wb.active

# set a title in the first row
#ws.cell(row=1, column=1, value='title')


"""
# go through all the folders' name to put their names in an Excel

for folder_name in os.listdir(folder_path):
    if os.path.isdir(os.path.join(folder_path, folder_name)):
        ws.append([folder_name])
        #ws.cell(row=row, column=1, value=folder_name)

"""

# go through all the folders' name to put their names in an Excel and build the each link
for row, folder_name in enumerate(os.listdir(folder_path), start=2):
    if os.path.isdir(os.path.join(folder_path, folder_name)):
        
        # write the folders' name into Excel
        ws.cell(row=row, column=1, value=folder_name)

        # build the hyperlink of each folder
        folder_full_path = os.path.join(folder_path, folder_name)
        # the path after "+" is for building a hyperlink for a specific file in the folders
        folder_full_path = folder_full_path # + "/models/model_normalized.obj" 
        hyperlink = Hyperlink(ref=f'A{row}', target=folder_full_path)
        ws.cell(row=row, column=1).hyperlink = hyperlink


# save the names to Excel (need to give a full path where the excel will be stored)
excel_file_path = r'C:\your own path\folders_names.xlsx'
wb.save(excel_file_path)

print("\n\n\nThe folders' names have put into an excel.\n")
print("Conversion complete!\n")